"""
Service Rapportage - Génération rapports PDF/Excel
"""

from io import BytesIO
from datetime import datetime, timedelta
from typing import Dict, Any, List
import json
import structlog

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    HAS_REPORTING = True
except ImportError:
    HAS_REPORTING = False
    print("⚠️ reportlab/openpyxl non disponibles - Rapportage limité")

logger = structlog.get_logger()


class ReportService:
    """Service génération rapports PDF/Excel"""

    def __init__(self):
        self.styles = getSampleStyleSheet() if HAS_REPORTING else None
        self._setup_custom_styles()

    def _setup_custom_styles(self):
        """Configurer styles personnalisés"""
        if not HAS_REPORTING:
            return

        self.title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1  # Center
        )

        self.subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=self.styles['Heading2'],
            fontSize=14,
            spaceAfter=20
        )

    def generate_farmer_report_pdf(self, user_data: Dict[str, Any], period_days: int = 30) -> bytes:
        """
        Générer rapport PDF agriculteur

        Args:
            user_data: Données utilisateur
            period_days: Période en jours

        Returns:
            PDF bytes
        """
        if not HAS_REPORTING:
            return self._fallback_pdf()

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []

        # Titre
        elements.append(Paragraph("RAPPORT AGRICOLE - AGROSMART", self.title_style))
        elements.append(Paragraph(f"Période: {period_days} jours", self.subtitle_style))
        elements.append(Spacer(1, 20))

        # Informations agriculteur
        farmer_info = [
            ["Agriculteur", user_data.get('name', 'N/A')],
            ["Région", user_data.get('region', 'N/A')],
            ["Date", datetime.now().strftime('%d/%m/%Y')]
        ]

        table = Table(farmer_info)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(table)
        elements.append(Spacer(1, 20))

        # Recommandations cultures
        elements.append(Paragraph("RECOMMANDATIONS CULTURES", self.subtitle_style))
        crops_data = user_data.get('crop_recommendations', [])
        if crops_data:
            crop_table_data = [["Culture", "Score", "Raison"]]
            for crop in crops_data[:5]:  # Top 5
                crop_table_data.append([
                    crop.get('crop', 'N/A'),
                    f"{crop.get('score', 0.0):.1f}",
                    crop.get('reason', 'N/A')
                ])

            crop_table = Table(crop_table_data)
            crop_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.green),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(crop_table)
        elements.append(Spacer(1, 20))

        # Prévisions météo
        elements.append(Paragraph("PRÉVISIONS MÉTÉO (7 jours)", self.subtitle_style))
        weather_data = user_data.get('weather_forecast', [])
        if weather_data:
            weather_table_data = [["Date", "Température", "Précipitations", "Risque"]]
            for day in weather_data[:7]:
                weather_table_data.append([
                    day.get('date', 'N/A'),
                    f"{day.get('temp', 0.0):.1f}°C" if day.get('temp') is not None else 'N/A',
                    f"{day.get('rain', 0.0):.1f}mm" if day.get('rain') is not None else 'N/A',
                    day.get('risk', 'N/A')
                ])

            weather_table = Table(weather_table_data)
            weather_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.blue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(weather_table)

        # Pied de page
        elements.append(Spacer(1, 50))
        elements.append(Paragraph("Rapport généré par AgroSmart - Plateforme IA Agricole", self.styles['Italic']))

        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()

    def generate_government_report_excel(self, region_data: Dict[str, Any]) -> bytes:
        """
        Générer rapport Excel gouvernement

        Args:
            region_data: Données régionales

        Returns:
            Excel bytes
        """
        if not HAS_REPORTING:
            return self._fallback_excel()

        wb = Workbook()
        ws = wb.active
        ws.title = "Rapport Régional"

        # En-têtes
        headers = ['Région', 'Agriculteurs', 'Superficie', 'Rendement Moyen', 'Problèmes Signalés', 'Recommandations']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Données
        row = 2
        for region, data in region_data.items():
            ws.cell(row=row, column=1, value=region)
            ws.cell(row=row, column=2, value=data.get('farmers_count', 0))
            ws.cell(row=row, column=3, value=data.get('area_ha', 0))
            ws.cell(row=row, column=4, value=data.get('avg_yield', 0))
            ws.cell(row=row, column=5, value=data.get('issues_count', 0))
            ws.cell(row=row, column=6, value=data.get('recommendations', ''))

            # Couleur conditionnelle
            if data.get('issues_count', 0) > 10:
                for col in range(1, 7):
                    ws.cell(row=row, column=col).fill = PatternFill(start_color="FFCCCC", fill_type="solid")

            row += 1

        # Ajuster colonnes
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 15

        # Sauvegarder
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def _fallback_pdf(self) -> bytes:
        """PDF fallback quand reportlab indisponible"""
        fallback_content = b"""
RAPPORT AGRICOLE - AGROSMART
============================

Mode Rapportage Limite
----------------------
Les librairies de rapportage ne sont pas installes.

Pour activer:
pip install reportlab openpyxl

Rapport genere le: """ + datetime.now().strftime('%Y-%m-%d %H:%M:%S').encode()

        return fallback_content

    def _fallback_excel(self) -> bytes:
        """Excel fallback"""
        return b"Mode Excel limite - installer openpyxl"


# Instance globale
report_service = ReportService()


def generate_pdf_report(user_data: Dict[str, Any], period_days: int = 30) -> bytes:
    """Fonction principale génération PDF"""
    return report_service.generate_farmer_report_pdf(user_data, period_days)


def generate_excel_report(region_data: Dict[str, Any]) -> bytes:
    """Fonction principale génération Excel"""
    return report_service.generate_government_report_excel(region_data)